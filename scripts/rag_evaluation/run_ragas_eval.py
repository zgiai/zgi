#!/usr/bin/env python3
"""Build a Ragas dataset from an Excel QA file, call ZGI RAG eval, and run Ragas."""

from __future__ import annotations

import argparse
import copy
import csv
import json
import os
import time
import urllib.error
import urllib.parse
import urllib.request
import warnings
import zipfile
from dataclasses import dataclass
from getpass import getpass
from io import BytesIO
from pathlib import Path
from typing import Any


DEFAULT_LIMIT = 0
DEFAULT_BACKEND_BATCH_SIZE = 10
DEFAULT_RAGAS_BATCH_SIZE = 10
DEFAULT_RAGAS_MAX_WORKERS = 8
DEFAULT_RAG_EVAL_TOP_K = 10
DEFAULT_RAG_EVAL_SCORE_THRESHOLD = 0.35
DEFAULT_BASE_URL = "http://127.0.0.1:2670/console/api"
DEFAULT_ALIYUN_RAGAS_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
DEFAULT_ALIYUN_RAGAS_LLM_MODEL = "qwen-plus"
DEFAULT_ALIYUN_RAGAS_EMBEDDING_MODEL = "text-embedding-v4"
SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = SCRIPT_DIR / "input"
MIDDLE_DIR = SCRIPT_DIR / "middle"
RESULT_DIR = SCRIPT_DIR / "result"
ENV_FILE = SCRIPT_DIR / ".env"
ENV_VALUES = {}
INPUT_EXTENSIONS = {".xls", ".xlsx", ".csv"}

QUESTION_HEADERS = {
    "question",
    "query",
    "user_input",
    "user input",
    "prompt",
    "问题",
    "用户问题",
    "提问",
}
ANSWER_HEADERS = {
    "answer",
    "reference",
    "ground_truth",
    "ground truth",
    "标准答案",
    "答案",
    "回答",
}


@dataclass
class QAItem:
    question: str
    reference: str


@dataclass
class RagasModelConfig:
    provider: str
    api_key: str
    base_url: str
    llm_model: str
    embedding_model: str
    enable_thinking: bool | None
    max_workers: int


class OpenAICompatibleRagasEmbeddings:
    def __init__(self, sync_client: Any, async_client: Any, model: str) -> None:
        self.sync_client = sync_client
        self.async_client = async_client
        self.model = model

    def embed_query(self, text: str) -> list[float]:
        return self.embed_documents([text])[0]

    def embed_documents(self, texts: list[str]) -> list[list[float]]:
        normalized = self._normalize_texts(texts)
        if not normalized:
            return []
        response = self.sync_client.embeddings.create(input=normalized, model=self.model)
        return [item.embedding for item in response.data]

    async def aembed_query(self, text: str) -> list[float]:
        return (await self.aembed_documents([text]))[0]

    async def aembed_documents(self, texts: list[str]) -> list[list[float]]:
        normalized = self._normalize_texts(texts)
        if not normalized:
            return []
        response = await self.async_client.embeddings.create(input=normalized, model=self.model)
        return [item.embedding for item in response.data]

    async def embed_text(self, text: str, is_async: bool = True) -> list[float]:
        if is_async:
            return await self.aembed_query(text)
        return self.embed_query(text)

    async def embed_texts(self, texts: list[str], is_async: bool = True) -> list[list[float]]:
        if is_async:
            return await self.aembed_documents(texts)
        return self.embed_documents(texts)

    @staticmethod
    def _normalize_texts(texts: list[str]) -> list[str]:
        return [str(text) for text in texts if str(text).strip()]


def main(output_platform: str = "") -> int:
    global ENV_VALUES
    ENV_VALUES = load_env_file(ENV_FILE)

    args = parse_args()
    input_path = choose_input_path(args.input)
    if not input_path.exists():
        raise SystemExit(f"input file does not exist: {input_path}")

    dataset_path, result_json_path, result_csv_path = output_paths_for_input(input_path, output_platform)
    dataset_rows: list[dict[str, Any]] | None = None
    if dataset_path.exists():
        if args.reuse_dataset:
            answer = "yes"
        elif args.recollect:
            answer = ""
        else:
            answer = input(
                f"Found existing Ragas dataset: {dataset_path}\n"
                "Use it directly for evaluation? Enter y/yes to reuse, or press Enter to collect backend data again: "
            ).strip().lower()
        if answer in {"y", "yes"}:
            dataset_rows = load_existing_dataset(dataset_path)
            print(f"loaded existing Ragas dataset: {dataset_path} ({len(dataset_rows)} rows)")

    email = args.email or env_value("ZGI_EMAIL") or input("ZGI email: ").strip()
    base_url = args.base_url.rstrip("/")
    ENV_VALUES["ZGI_BASE_URL"] = base_url
    ENV_VALUES["ZGI_EMAIL"] = email

    if dataset_rows is None:
        qa_items = read_qa_items(input_path, args.limit)
        if not qa_items:
            raise SystemExit("no QA rows found in input file")

        token = get_cached_token(base_url, email)
        if not token:
            token = interactive_login(base_url, email, args.password)
            write_cached_token(base_url, email, token)

        knowledge_base_name = args.knowledge_base_name or env_value("ZGI_KNOWLEDGE_BASE_NAME")
        if not knowledge_base_name:
            knowledge_base_name = input("Knowledge base name: ").strip()
        if not knowledge_base_name:
            raise SystemExit("knowledge base name is required")
        ENV_VALUES["ZGI_KNOWLEDGE_BASE_NAME"] = knowledge_base_name
        write_env_file(ENV_FILE, ENV_VALUES)
        top_k, score_threshold = resolve_retrieval_eval_params(args)

        questions = [item.question for item in qa_items]
        try:
            eval_items = call_rag_evaluation(
                base_url,
                token,
                knowledge_base_name,
                questions,
                top_k,
                score_threshold,
                args.retrieval_mode,
                args.model,
                args.backend_batch_size,
            )
        except HTTPStatusError as exc:
            if exc.status == 401:
                print("cached token is invalid or expired; please log in again.")
                token = interactive_login(base_url, email, args.password)
                write_cached_token(base_url, email, token)
                eval_items = call_rag_evaluation(
                    base_url,
                    token,
                    knowledge_base_name,
                    questions,
                    top_k,
                    score_threshold,
                    args.retrieval_mode,
                    args.model,
                    args.backend_batch_size,
                )
            else:
                raise
        except urllib.error.URLError as exc:
            raise SystemExit(f"cannot connect to {base_url}: {exc}") from exc

        dataset_rows = build_ragas_rows(qa_items, eval_items)
        write_json(dataset_path, dataset_rows)
        print(f"saved Ragas dataset: {dataset_path}")

    ragas_model_config = build_ragas_model_config(args)
    remember_ragas_model_config(ragas_model_config)
    ENV_VALUES["RAGAS_BATCH_SIZE"] = str(args.ragas_batch_size)
    write_env_file(ENV_FILE, ENV_VALUES)
    results = run_ragas(dataset_rows, ragas_model_config, args.ragas_batch_size, args.ragas_limit)
    write_ragas_outputs(results, result_json_path, result_csv_path)
    print(f"saved Ragas result JSON: {result_json_path}")
    print(f"saved Ragas result CSV: {result_csv_path}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run ZGI RAG evaluation and Ragas metrics for an input QA file.")
    parser.add_argument("--input", default="", help="Optional input file path. If omitted, choose from scripts/rag_evaluation/input.")
    parser.add_argument("--limit", type=int, default=DEFAULT_LIMIT, help="Number of QA rows to evaluate. 0 means all rows. Default: %(default)s")
    parser.add_argument("--base-url", default=env_value("ZGI_BASE_URL", DEFAULT_BASE_URL), help="API v1 base URL.")
    parser.add_argument("--email", default=env_value("ZGI_EMAIL"), help="Login email. Can also use ZGI_EMAIL or .env.")
    parser.add_argument("--password", default=env_value("ZGI_PASSWORD"), help="Login password. Can also use ZGI_PASSWORD or .env.")
    parser.add_argument("--knowledge-base-name", default=env_value("ZGI_KNOWLEDGE_BASE_NAME"), help="Knowledge base name.")
    parser.add_argument("--top-k", type=int, default=None, help="Retrieval top_k passed to the backend. If omitted, prompt and save to .env.")
    parser.add_argument("--score-threshold", type=float, default=None, help="Retrieval score threshold passed to the backend. If omitted, prompt and save to .env.")
    parser.add_argument("--backend-batch-size", type=int, default=DEFAULT_BACKEND_BATCH_SIZE, help="Questions per backend request. Default: %(default)s")
    parser.add_argument("--ragas-batch-size", type=int, default=int_env_value("RAGAS_BATCH_SIZE", DEFAULT_RAGAS_BATCH_SIZE), help="Rows per Ragas evaluation batch after all backend data is collected. Default: %(default)s")
    parser.add_argument("--ragas-limit", type=int, default=0, help="Limit rows sent to Ragas after filtering successful backend rows. 0 means no limit.")
    parser.add_argument("--retrieval-mode", default="hybrid", choices=["hybrid", "vector", "graph"], help="Retrieval mode.")
    parser.add_argument("--model", default="", help="Optional generation model name for the backend.")
    parser.add_argument("--ragas-provider", default=env_value("RAGAS_PROVIDER", "auto"), choices=["auto", "aliyun", "openai"], help="Ragas judge provider.")
    parser.add_argument("--ragas-api-key", default=ragas_env_value("RAGAS_API_KEY", "ALIYUN_API_KEY", "DASHSCOPE_API_KEY", "OPENAI_API_KEY"), help="Ragas judge API key.")
    parser.add_argument("--ragas-base-url", default=ragas_env_value("RAGAS_BASE_URL", "ALIYUN_BASE_URL", "DASHSCOPE_BASE_URL", "OPENAI_BASE_URL"), help="OpenAI-compatible base URL for Ragas.")
    parser.add_argument("--ragas-llm-model", default=ragas_env_value("RAGAS_LLM_MODEL", "ALIYUN_LLM_MODEL", "DASHSCOPE_LLM_MODEL"), help="Chat model used by Ragas.")
    parser.add_argument("--ragas-embedding-model", default=ragas_env_value("RAGAS_EMBEDDING_MODEL", "ALIYUN_EMBEDDING_MODEL", "DASHSCOPE_EMBEDDING_MODEL"), help="Embedding model used by Ragas.")
    parser.add_argument("--ragas-enable-thinking", default=ragas_env_value("RAGAS_ENABLE_THINKING", "ALIYUN_ENABLE_THINKING", "DASHSCOPE_ENABLE_THINKING"), help="Enable DashScope thinking mode for Ragas judge LLM. true/false.")
    parser.add_argument("--ragas-max-workers", type=int, default=int_env_value("RAGAS_MAX_WORKERS", DEFAULT_RAGAS_MAX_WORKERS), help="Ragas concurrent workers. Default: %(default)s")
    dataset_group = parser.add_mutually_exclusive_group()
    dataset_group.add_argument("--reuse-dataset", action="store_true", help="Reuse the existing platform dataset without collecting backend data.")
    dataset_group.add_argument("--recollect", action="store_true", help="Ignore an existing platform dataset and recollect backend data.")
    return parser.parse_args()


def load_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except FileNotFoundError:
        return values
    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            values[key] = value
    return values


def env_value(key: str, default: str = "") -> str:
    return os.getenv(key) or ENV_VALUES.get(key, default)


def ragas_env_value(*keys: str) -> str:
    for key in keys:
        value = env_value(key)
        if value:
            return value
    return ""


def int_env_value(key: str, default: int) -> int:
    value = env_value(key)
    if not value:
        return default
    try:
        return int(value)
    except ValueError as exc:
        raise SystemExit(f"{key} must be an integer, got: {value}") from exc


def float_env_value(key: str, default: float) -> float:
    value = env_value(key)
    if not value:
        return default
    try:
        return float(value)
    except ValueError as exc:
        raise SystemExit(f"{key} must be a number, got: {value}") from exc


def resolve_retrieval_eval_params(args: argparse.Namespace) -> tuple[int, float]:
    if args.top_k is not None and args.score_threshold is not None:
        top_k = require_top_k(args.top_k)
        score_threshold = require_score_threshold(args.score_threshold)
        save_retrieval_eval_params(top_k, score_threshold)
        return top_k, score_threshold

    if args.top_k is not None or args.score_threshold is not None:
        top_k = require_top_k(args.top_k) if args.top_k is not None else prompt_top_k(default_saved_top_k())
        score_threshold = (
            require_score_threshold(args.score_threshold)
            if args.score_threshold is not None
            else prompt_score_threshold(default_saved_score_threshold())
        )
        save_retrieval_eval_params(top_k, score_threshold)
        return top_k, score_threshold

    saved_top_k = env_value("ZGI_RAG_EVAL_TOP_K")
    saved_score_threshold = env_value("ZGI_RAG_EVAL_SCORE_THRESHOLD")
    if saved_top_k and saved_score_threshold:
        top_k = validate_top_k(int_env_value("ZGI_RAG_EVAL_TOP_K", DEFAULT_RAG_EVAL_TOP_K))
        score_threshold = validate_score_threshold(
            float_env_value("ZGI_RAG_EVAL_SCORE_THRESHOLD", DEFAULT_RAG_EVAL_SCORE_THRESHOLD)
        )
        answer = input(
            f"Saved retrieval defaults: top_k={top_k}, score_threshold={score_threshold}. "
            "Press Enter to use them, or enter r to reconfigure: "
        ).strip().lower()
        if answer != "r":
            return top_k, score_threshold

    top_k = prompt_top_k(default_saved_top_k())
    score_threshold = prompt_score_threshold(default_saved_score_threshold())
    save_retrieval_eval_params(top_k, score_threshold)
    return top_k, score_threshold


def default_saved_top_k() -> int:
    return int_env_value("ZGI_RAG_EVAL_TOP_K", DEFAULT_RAG_EVAL_TOP_K)


def default_saved_score_threshold() -> float:
    return float_env_value("ZGI_RAG_EVAL_SCORE_THRESHOLD", DEFAULT_RAG_EVAL_SCORE_THRESHOLD)


def prompt_top_k(default: int) -> int:
    while True:
        raw = input(f"Retrieval top_k [{default}]: ").strip()
        value = default if raw == "" else raw
        try:
            return validate_top_k(int(value))
        except ValueError:
            print("top_k must be an integer between 1 and 20.")


def prompt_score_threshold(default: float) -> float:
    while True:
        raw = input(f"Retrieval score_threshold [{default}]: ").strip()
        value = default if raw == "" else raw
        try:
            return validate_score_threshold(float(value))
        except ValueError:
            print("score_threshold must be a number between 0 and 1.")


def validate_top_k(value: int) -> int:
    if value < 1 or value > 20:
        raise ValueError("top_k must be between 1 and 20")
    return value


def validate_score_threshold(value: float) -> float:
    if value < 0 or value > 1:
        raise ValueError("score_threshold must be between 0 and 1")
    return value


def require_top_k(value: int) -> int:
    try:
        return validate_top_k(value)
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc


def require_score_threshold(value: float) -> float:
    try:
        return validate_score_threshold(value)
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc


def save_retrieval_eval_params(top_k: int, score_threshold: float) -> None:
    ENV_VALUES["ZGI_RAG_EVAL_TOP_K"] = str(top_k)
    ENV_VALUES["ZGI_RAG_EVAL_SCORE_THRESHOLD"] = str(score_threshold)
    write_env_file(ENV_FILE, ENV_VALUES)


def optional_bool(value: str) -> bool | None:
    normalized = (value or "").strip().lower()
    if not normalized:
        return None
    if normalized in {"1", "true", "yes", "y", "on"}:
        return True
    if normalized in {"0", "false", "no", "n", "off"}:
        return False
    raise SystemExit(f"invalid boolean value: {value}; expected true or false")


def choose_input_path(input_arg: str) -> Path:
    if input_arg:
        return Path(input_arg).expanduser().resolve()

    input_files = available_input_files(INPUT_DIR)
    if not input_files:
        raise SystemExit(
            f"no input files found in {INPUT_DIR}; supported extensions: "
            f"{', '.join(sorted(INPUT_EXTENSIONS))}"
        )

    print(f"Available input files in {INPUT_DIR}:")
    for idx, path in enumerate(input_files, start=1):
        print(f"  {idx}. {path.name}")

    while True:
        choice = input("Select input file number: ").strip()
        try:
            selected = int(choice)
        except ValueError:
            print("please enter a number from the list")
            continue
        if 1 <= selected <= len(input_files):
            input_path = input_files[selected - 1]
            print(f"selected input file: {input_path}")
            return input_path
        print(f"please enter a number between 1 and {len(input_files)}")


def available_input_files(input_dir: Path) -> list[Path]:
    if not input_dir.exists():
        return []
    return sorted(
        path
        for path in input_dir.iterdir()
        if path.is_file() and not path.name.startswith(("~$", ".")) and path.suffix.lower() in INPUT_EXTENSIONS
    )


def output_paths_for_input(input_path: Path, platform: str = "") -> tuple[Path, Path, Path]:
    MIDDLE_DIR.mkdir(parents=True, exist_ok=True)
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    dataset_prefix = MIDDLE_DIR / input_path.with_suffix("").name
    result_prefix = RESULT_DIR / input_path.with_suffix("").name
    platform_suffix = f".{platform.strip().lower()}" if platform.strip() else ""
    return (
        dataset_prefix.with_name(dataset_prefix.name + platform_suffix + ".ragas.dataset.json"),
        result_prefix.with_name(result_prefix.name + platform_suffix + ".ragas.results.json"),
        result_prefix.with_name(result_prefix.name + platform_suffix + ".ragas.results.csv"),
    )


def load_existing_dataset(path: Path) -> list[dict[str, Any]]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SystemExit(f"existing Ragas dataset is not valid JSON: {path}") from exc
    if not isinstance(data, list):
        raise SystemExit(f"existing Ragas dataset must be a JSON list: {path}")
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(data, start=1):
        if not isinstance(row, dict):
            raise SystemExit(f"existing Ragas dataset row {idx} is not a JSON object: {path}")
        rows.append(row)
    return rows


def build_ragas_model_config(args: argparse.Namespace) -> RagasModelConfig:
    provider = (args.ragas_provider or "auto").strip().lower()
    api_key = (args.ragas_api_key or "").strip()
    base_url = (args.ragas_base_url or "").strip().rstrip("/")

    if provider == "auto":
        if env_value("ALIYUN_API_KEY") or env_value("DASHSCOPE_API_KEY") or "dashscope.aliyuncs.com" in base_url:
            provider = "aliyun"
        else:
            provider = "openai"

    if provider == "aliyun":
        base_url = base_url or DEFAULT_ALIYUN_RAGAS_BASE_URL
        llm_model = (args.ragas_llm_model or DEFAULT_ALIYUN_RAGAS_LLM_MODEL).strip()
        embedding_model = (args.ragas_embedding_model or DEFAULT_ALIYUN_RAGAS_EMBEDDING_MODEL).strip()
        enable_thinking = optional_bool(args.ragas_enable_thinking)
        if enable_thinking is None:
            enable_thinking = False
    else:
        llm_model = (args.ragas_llm_model or "gpt-4o-mini").strip()
        embedding_model = (args.ragas_embedding_model or "text-embedding-3-small").strip()
        enable_thinking = None

    if not api_key:
        if provider == "aliyun":
            raise SystemExit(
                "Ragas requires an Aliyun DashScope API key. Set RAGAS_API_KEY, ALIYUN_API_KEY, "
                "or DASHSCOPE_API_KEY in scripts/rag_evaluation/.env."
            )
        raise SystemExit("Ragas requires an API key. Set RAGAS_API_KEY or OPENAI_API_KEY in scripts/rag_evaluation/.env.")
    if not llm_model:
        raise SystemExit("Ragas LLM model is required. Set RAGAS_LLM_MODEL in scripts/rag_evaluation/.env.")
    if not embedding_model:
        raise SystemExit("Ragas embedding model is required. Set RAGAS_EMBEDDING_MODEL in scripts/rag_evaluation/.env.")
    if args.ragas_max_workers < 1:
        raise SystemExit("RAGAS_MAX_WORKERS must be >= 1.")

    return RagasModelConfig(
        provider=provider,
        api_key=api_key,
        base_url=base_url,
        llm_model=llm_model,
        embedding_model=embedding_model,
        enable_thinking=enable_thinking,
        max_workers=args.ragas_max_workers,
    )


def remember_ragas_model_config(config: RagasModelConfig) -> None:
    ENV_VALUES["RAGAS_PROVIDER"] = config.provider
    if config.base_url:
        ENV_VALUES["RAGAS_BASE_URL"] = config.base_url
    ENV_VALUES["RAGAS_LLM_MODEL"] = config.llm_model
    ENV_VALUES["RAGAS_EMBEDDING_MODEL"] = config.embedding_model
    if config.enable_thinking is not None:
        ENV_VALUES["RAGAS_ENABLE_THINKING"] = "true" if config.enable_thinking else "false"
    ENV_VALUES["RAGAS_MAX_WORKERS"] = str(config.max_workers)
    ENV_VALUES["RAGAS_API_KEY"] = config.api_key
    write_env_file(ENV_FILE, ENV_VALUES)


def read_qa_items(path: Path, limit: int) -> list[QAItem]:
    rows = read_input_rows(path)
    if not rows:
        return []

    start_idx = detect_start_row(rows)
    items: list[QAItem] = []
    for row in rows[start_idx:]:
        question = clean_cell(row[0] if len(row) > 0 else None)
        reference = clean_cell(row[1] if len(row) > 1 else None)
        if not question or not reference:
            continue
        items.append(QAItem(question=question, reference=reference))
        if limit > 0 and len(items) >= limit:
            break
    return items


def read_input_rows(path: Path) -> list[list[Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xls", ".xlsx"}:
        return read_excel_rows(path)
    if suffix == ".csv":
        return read_csv_rows(path)
    raise SystemExit(f"unsupported input file type: {path.suffix}")


def read_csv_rows(path: Path) -> list[list[Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [row for row in csv.reader(f)]


def read_excel_rows(path: Path) -> list[list[Any]]:
    with path.open("rb") as f:
        data = f.read()
    if zipfile.is_zipfile(BytesIO(data)):
        return read_openpyxl_rows(data)
    return read_legacy_xls_rows(path)


def read_openpyxl_rows(data: bytes) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required to read xlsx-format Excel files. Install it with: pip install openpyxl") from exc

    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def read_legacy_xls_rows(path: Path) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise SystemExit("xlrd is required to read legacy .xls files. Install it with: pip install xlrd") from exc

    workbook = xlrd.open_workbook(str(path))
    sheet = workbook.sheet_by_index(0)
    return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]


def detect_start_row(rows: list[list[Any]]) -> int:
    first = [clean_header(cell) for cell in rows[0]]
    if len(first) >= 2 and first[0] in QUESTION_HEADERS and first[1] in ANSWER_HEADERS:
        return 1
    return 0


def clean_header(value: Any) -> str:
    return clean_cell(value).lower().replace("\ufeff", "")


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        text = text[1:-1].strip()
    return text


def get_cached_token(base_url: str, email: str) -> str:
    if ENV_VALUES.get("ZGI_BASE_URL", base_url).rstrip("/") != base_url:
        return ""
    if ENV_VALUES.get("ZGI_EMAIL", email) != email:
        return ""
    return ENV_VALUES.get("ZGI_ACCESS_TOKEN", "")


def write_cached_token(base_url: str, email: str, token: str) -> None:
    ENV_VALUES.update(
        {
            "ZGI_BASE_URL": base_url,
            "ZGI_EMAIL": email,
            "ZGI_ACCESS_TOKEN": token,
            "ZGI_TOKEN_CACHED_AT": str(int(time.time())),
        }
    )
    if "ZGI_KNOWLEDGE_BASE_NAME" not in ENV_VALUES and env_value("ZGI_KNOWLEDGE_BASE_NAME"):
        ENV_VALUES["ZGI_KNOWLEDGE_BASE_NAME"] = env_value("ZGI_KNOWLEDGE_BASE_NAME")
    write_env_file(ENV_FILE, ENV_VALUES)
    try:
        ENV_FILE.chmod(0o600)
    except OSError:
        pass


def write_env_file(path: Path, values: dict[str, str]) -> None:
    ordered_keys = [
        "DIFY_BASE_URL",
        "DIFY_API_KEY",
        "DIFY_USER_PREFIX",
        "DIFY_RESPONSE_MODE",
        "ZGI_BASE_URL",
        "ZGI_EMAIL",
        "ZGI_KNOWLEDGE_BASE_NAME",
        "ZGI_RAG_EVAL_TOP_K",
        "ZGI_RAG_EVAL_SCORE_THRESHOLD",
        "ZGI_ACCESS_TOKEN",
        "ZGI_TOKEN_CACHED_AT",
        "RAGAS_PROVIDER",
        "RAGAS_BASE_URL",
        "RAGAS_LLM_MODEL",
        "RAGAS_EMBEDDING_MODEL",
        "RAGAS_ENABLE_THINKING",
        "RAGAS_BATCH_SIZE",
        "RAGAS_MAX_WORKERS",
        "RAGAS_API_KEY",
    ]
    lines = []
    for key in ordered_keys:
        value = values.get(key, "")
        if value:
            lines.append(f"{key}={shell_quote_env(value)}")
    for key in sorted(values):
        if key not in ordered_keys and values[key]:
            lines.append(f"{key}={shell_quote_env(values[key])}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    try:
        path.chmod(0o600)
    except OSError:
        pass


def shell_quote_env(value: str) -> str:
    escaped = value.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def interactive_login(base_url: str, email: str, password_arg: str) -> str:
    password = password_arg or env_value("ZGI_PASSWORD") or getpass("ZGI password: ")
    try:
        token = login(base_url, email, password)
    except HTTPStatusError as exc:
        raise SystemExit(f"login failed with HTTP {exc.status}: {exc.body}") from exc
    except urllib.error.URLError as exc:
        raise SystemExit(f"cannot connect to {base_url}: {exc}") from exc
    print(f"login succeeded; token cached in {ENV_FILE}")
    return token


def login(base_url: str, email: str, password: str) -> str:
    payload = {"email": email, "password": password, "remember_me": True}
    data = post_json(f"{base_url}/login", payload, token="")
    token = extract_access_token(data)
    if not token:
        raise SystemExit("login succeeded but no access token was found in response")
    return token


def call_rag_evaluation(
    base_url: str,
    token: str,
    knowledge_base_name: str,
    questions: list[str],
    top_k: int,
    score_threshold: float,
    retrieval_mode: str,
    model: str,
    batch_size: int,
) -> list[dict[str, Any]]:
    batch_size = normalize_batch_size(batch_size, DEFAULT_BACKEND_BATCH_SIZE)
    total = len(questions)
    all_items: list[dict[str, Any]] = []
    print(
        f"collecting backend RAG data: {total} questions, batch_size={batch_size}, "
        f"top_k={top_k}, score_threshold={score_threshold}",
        flush=True,
    )
    for start in range(0, total, batch_size):
        batch = questions[start : start + batch_size]
        end = start + len(batch)
        print(f"backend batch {start + 1}-{end}/{total} started", flush=True)
        items = call_rag_evaluation_batch(
            base_url,
            token,
            knowledge_base_name,
            batch,
            top_k,
            score_threshold,
            retrieval_mode,
            model,
        )
        if len(items) != len(batch):
            raise SystemExit(f"backend batch {start + 1}-{end} returned {len(items)} rows for {len(batch)} questions")
        all_items.extend(items)
        print(f"backend batch {start + 1}-{end}/{total} finished", flush=True)
    print(f"backend RAG data collection finished: {len(all_items)}/{total}", flush=True)
    return all_items


def call_rag_evaluation_batch(
    base_url: str,
    token: str,
    knowledge_base_name: str,
    questions: list[str],
    top_k: int,
    score_threshold: float,
    retrieval_mode: str,
    model: str,
) -> list[dict[str, Any]]:
    payload = {
        "knowledge_base_name": knowledge_base_name,
        "user_inputs": questions,
        "top_k": top_k,
        "score_threshold": score_threshold,
        "retrieval_mode": retrieval_mode,
    }
    if model:
        payload["model"] = model
    data = post_json(f"{base_url}/rag-evaluation/batch", payload, token=token)
    body = data.get("data", data)
    items = body.get("data") if isinstance(body, dict) else None
    if not isinstance(items, list):
        raise SystemExit("rag-evaluation response does not contain data.data list")
    return items


def post_json(url: str, payload: dict[str, Any], token: str) -> dict[str, Any]:
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    request = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({})) if is_local_url(url) else urllib.request.build_opener()
    try:
        with opener.open(request, timeout=300) as response:
            raw = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise HTTPStatusError(exc.code, detail) from exc
    return json.loads(raw)


def is_local_url(url: str) -> bool:
    hostname = urllib.parse.urlparse(url).hostname or ""
    return hostname in {"localhost", "127.0.0.1", "::1"}


def extract_access_token(data: dict[str, Any]) -> str:
    candidates = [
        data.get("access_token"),
        data.get("data", {}).get("access_token") if isinstance(data.get("data"), dict) else None,
        data.get("data", {}).get("data", {}).get("access_token") if isinstance(data.get("data"), dict) and isinstance(data.get("data", {}).get("data"), dict) else None,
    ]
    for candidate in candidates:
        if isinstance(candidate, str) and candidate:
            return candidate
    return ""


def build_ragas_rows(qa_items: list[QAItem], eval_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if len(eval_items) != len(qa_items):
        raise SystemExit(f"backend returned {len(eval_items)} rows for {len(qa_items)} questions")

    rows: list[dict[str, Any]] = []
    for sample_id, (qa, result) in enumerate(zip(qa_items, eval_items), start=1):
        contexts = result.get("retrieved_contexts") or []
        if not isinstance(contexts, list):
            contexts = []
        rows.append(
            {
                "sample_id": sample_id,
                "platform": "zgi",
                "user_input": qa.question,
                "response": str(result.get("response") or ""),
                "retrieved_contexts": [str(ctx) for ctx in contexts],
                "reference": qa.reference,
                "status": str(result.get("status") or ""),
                "error": str(result.get("error") or ""),
            }
        )
    return rows


def run_ragas(dataset_rows: list[dict[str, Any]], config: RagasModelConfig, batch_size: int, ragas_limit: int = 0) -> Any:
    try:
        from ragas import EvaluationDataset, evaluate
        from ragas.llms import llm_factory
        from openai import AsyncOpenAI, OpenAI
    except ImportError as exc:
        raise SystemExit(
            "ragas and its dataset dependencies are required. Install them in your Python environment with: "
            "pip install ragas datasets openai"
        ) from exc

    skipped_rows = [
        (idx, row)
        for idx, row in enumerate(dataset_rows, start=1)
        if not row["response"] or row.get("error")
    ]
    if skipped_rows:
        print(f"skipping {len(skipped_rows)} rows without response or with backend errors before Ragas evaluation:", flush=True)
        for idx, row in skipped_rows[:10]:
            reason = row.get("error") or "empty response"
            print(f"  row {idx}: {reason}", flush=True)
        if len(skipped_rows) > 10:
            print(f"  ... {len(skipped_rows) - 10} more skipped rows", flush=True)

    empty_context_rows = [
        idx
        for idx, row in enumerate(dataset_rows, start=1)
        if row["response"] and not row.get("error") and not row["retrieved_contexts"]
    ]
    if empty_context_rows:
        print(
            "including "
            f"{len(empty_context_rows)} rows with empty retrieved_contexts in Ragas evaluation "
            f"(rows: {', '.join(str(idx) for idx in empty_context_rows[:20])})",
            flush=True,
        )

    eligible_rows = [
        row
        for row in dataset_rows
        if row["response"] and not row.get("error")
    ]
    metric_rows = [
        {
            "user_input": row["user_input"],
            "response": row["response"],
            "retrieved_contexts": row["retrieved_contexts"],
            "reference": row["reference"],
        }
        for row in eligible_rows
    ]
    if not metric_rows:
        raise SystemExit("no rows with responses are available for Ragas evaluation")
    if ragas_limit > 0:
        metric_rows = metric_rows[:ragas_limit]
        eligible_rows = eligible_rows[:ragas_limit]

    llm_client = AsyncOpenAI(
        api_key=config.api_key,
        base_url=config.base_url or None,
        timeout=600,
        max_retries=2,
    )
    embedding_client = AsyncOpenAI(
        api_key=config.api_key,
        base_url=config.base_url or None,
        timeout=600,
        max_retries=2,
    )
    sync_embedding_client = OpenAI(
        api_key=config.api_key,
        base_url=config.base_url or None,
        timeout=600,
        max_retries=2,
    )
    llm_kwargs: dict[str, Any] = {"temperature": 0, "max_tokens": 4096}
    if config.provider == "aliyun" and config.enable_thinking is not None:
        llm_kwargs["extra_body"] = {"enable_thinking": config.enable_thinking}
    llm = llm_factory(config.llm_model, provider="openai", client=llm_client, **llm_kwargs)
    embeddings = OpenAICompatibleRagasEmbeddings(sync_embedding_client, embedding_client, config.embedding_model)
    metrics = build_ragas_metrics()
    print(
        "running Ragas with "
        f"provider={config.provider}, llm_model={config.llm_model}, "
        f"embedding_model={config.embedding_model}, enable_thinking={config.enable_thinking}, "
        f"max_workers={config.max_workers}"
    )
    batch_size = normalize_batch_size(batch_size, DEFAULT_RAGAS_BATCH_SIZE)
    total = len(metric_rows)
    print(f"Ragas evaluation started after dataset collection: {total} rows, batch_size={batch_size}", flush=True)
    result_rows: list[dict[str, Any]] = []
    ragas_started = time.perf_counter()
    for start in range(0, total, batch_size):
        batch = metric_rows[start : start + batch_size]
        end = start + len(batch)
        batch_started = time.perf_counter()
        print(f"Ragas batch {start + 1}-{end}/{total} started", flush=True)
        batch_result = evaluate_ragas_batch(batch, metrics, llm, embeddings, config.max_workers)
        result_rows.extend(ragas_result_to_rows(batch_result))
        batch_elapsed = time.perf_counter() - batch_started
        print(f"Ragas batch {start + 1}-{end}/{total} finished in {batch_elapsed:.1f}s", flush=True)
    ragas_elapsed = time.perf_counter() - ragas_started
    if len(result_rows) != len(eligible_rows):
        raise SystemExit(
            f"Ragas returned {len(result_rows)} rows for {len(eligible_rows)} evaluated rows; cannot preserve sample identity"
        )
    for fallback_id, (result_row, source_row) in enumerate(zip(result_rows, eligible_rows), start=1):
        result_row["sample_id"] = source_row.get("sample_id", fallback_id)
        if source_row.get("platform"):
            result_row["platform"] = source_row["platform"]
    print(f"Ragas evaluation finished: {len(result_rows)}/{total} in {ragas_elapsed:.1f}s", flush=True)
    return result_rows


def build_ragas_metrics() -> list[Any]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", DeprecationWarning)
        from ragas.metrics import answer_correctness, answer_relevancy, context_precision, context_recall, faithfulness

    metrics = copy.deepcopy([faithfulness, answer_relevancy, context_precision, context_recall, answer_correctness])
    for metric in metrics:
        if getattr(metric, "name", "") == "answer_relevancy" and hasattr(metric, "strictness"):
            metric.strictness = 1
    return metrics


def evaluate_ragas_batch(metric_rows: list[dict[str, Any]], metrics: list[Any], llm: Any, embeddings: Any, max_workers: int) -> Any:
    from ragas import EvaluationDataset, evaluate
    from ragas.run_config import RunConfig

    run_config = RunConfig(timeout=600, max_retries=3, max_wait=30, max_workers=max_workers)

    try:
        dataset = EvaluationDataset.from_list(metric_rows)
        return evaluate(dataset=dataset, metrics=metrics, llm=llm, embeddings=embeddings, run_config=run_config, raise_exceptions=True)
    except TypeError:
        from datasets import Dataset

        return evaluate(Dataset.from_list(metric_rows), metrics=metrics, llm=llm, embeddings=embeddings, run_config=run_config, raise_exceptions=True)


def ragas_result_to_rows(results: Any) -> list[dict[str, Any]]:
    if hasattr(results, "to_pandas"):
        return results.to_pandas().to_dict(orient="records")
    if isinstance(results, list):
        return results
    if isinstance(results, dict):
        return [results]
    if hasattr(results, "to_dict"):
        data = results.to_dict()
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            return [data]
    return [{"result": str(results)}]


def normalize_batch_size(value: int, default: int) -> int:
    if value <= 0:
        return default
    return value


def write_ragas_outputs(results: Any, json_path: Path, csv_path: Path) -> None:
    if hasattr(results, "to_pandas"):
        frame = results.to_pandas()
        json_path.write_text(frame.to_json(orient="records", force_ascii=False, indent=2), encoding="utf-8")
        frame.to_csv(csv_path, index=False)
        return

    data = results.to_dict() if hasattr(results, "to_dict") else results
    write_json(json_path, data)
    if isinstance(data, list):
        write_csv(csv_path, data)
    elif isinstance(data, dict):
        write_csv(csv_path, [data])


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = sorted({key for row in rows for key in row.keys()})
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


class HTTPStatusError(RuntimeError):
    def __init__(self, status: int, body: str) -> None:
        super().__init__(f"HTTP {status}: {body}")
        self.status = status
        self.body = body


if __name__ == "__main__":
    raise SystemExit(main())
