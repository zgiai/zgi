import json
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter


PRICE_FILE = Path("references/prices.json")
OUTPUT_DIR = Path("artifacts")


def normalize(value):
    return re.sub(r"\s+", "", str(value or "")).lower()


def cell_text(cell):
    return "" if cell.value is None else str(cell.value).strip()


def compact_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def safe_filename_part(value):
    text = compact_text(value)
    text = re.sub(r'[\\/:*?"<>|]+', "-", text)
    text = re.sub(r"\s+", "", text)
    text = text.strip(" ._-")
    return text


def merged_range_for_cell(sheet, row, col):
    for cell_range in sheet.merged_cells.ranges:
        if cell_range.min_row <= row <= cell_range.max_row and cell_range.min_col <= col <= cell_range.max_col:
            return cell_range
    return None


def writable_cell_to_right(sheet, row, col):
    label_range = merged_range_for_cell(sheet, row, col)
    start_col = label_range.max_col + 1 if label_range else col + 1
    scan_limit = max(sheet.max_column + 1, start_col)
    candidate_col = start_col
    while candidate_col <= scan_limit:
        target_range = merged_range_for_cell(sheet, row, candidate_col)
        if not target_range:
            return sheet.cell(row=row, column=candidate_col)
        if target_range.min_row == row and target_range.min_col >= start_col:
            return sheet.cell(row=target_range.min_row, column=target_range.min_col)
        candidate_col = target_range.max_col + 1
    return sheet.cell(row=row, column=start_col)


def sheet_for_update(wb, sheet_name):
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb[wb.sheetnames[0]]


def writable_cell_for_coordinate(sheet, coordinate):
    row, col = coordinate_to_tuple(str(coordinate).upper())
    cell_range = merged_range_for_cell(sheet, row, col)
    if cell_range:
        return sheet.cell(row=cell_range.min_row, column=cell_range.min_col)
    return sheet.cell(row=row, column=col)


def excel_value(value, value_type=""):
    if value_type == "number":
        parsed = number(value)
        return parsed if parsed is not None else value
    if value_type == "string":
        return "" if value is None else str(value)
    return value


def inspect_workbook(wb, args):
    max_rows = int(args.get("inspect_max_rows") or 80)
    max_cols = int(args.get("inspect_max_cols") or 30)
    max_cells = int(args.get("inspect_max_cells") or 800)
    sheets = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        cells = []
        for row in range(1, min(sheet.max_row, max_rows) + 1):
            for col in range(1, min(sheet.max_column, max_cols) + 1):
                cell = sheet.cell(row=row, column=col)
                text = cell_text(cell)
                if not text:
                    continue
                item = {
                    "cell": f"{get_column_letter(col)}{row}",
                    "row": row,
                    "column": col,
                    "value": text,
                }
                merged = merged_range_for_cell(sheet, row, col)
                if merged:
                    item["merged_range"] = str(merged)
                    item["merged_anchor"] = f"{get_column_letter(merged.min_col)}{merged.min_row}"
                cells.append(item)
                if len(cells) >= max_cells:
                    break
            if len(cells) >= max_cells:
                break
        sheets.append({
            "name": sheet_name,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "cells": cells,
        })
    return {"sheets": sheets}


def apply_cell_updates(wb, updates, default_sheet_name):
    applied = []
    for update in updates or []:
        if not isinstance(update, dict):
            continue
        sheet = sheet_for_update(wb, update.get("sheet_name") or default_sheet_name)
        right_of = update.get("right_of_cell")
        if right_of:
            row, col = coordinate_to_tuple(str(right_of).upper())
            cell = writable_cell_to_right(sheet, row, col)
            coordinate = str(right_of).upper()
        else:
            cell = None
            coordinate = update.get("cell")
        if not coordinate and update.get("row") and update.get("column"):
            coordinate = f"{get_column_letter(int(update['column']))}{int(update['row'])}"
        if not coordinate:
            continue
        if cell is None:
            cell = writable_cell_for_coordinate(sheet, coordinate)
        value = update.get("value")
        if update.get("formula"):
            formula = str(update.get("formula")).strip()
            value = formula if formula.startswith("=") else "=" + formula
        else:
            value = excel_value(value, str(update.get("value_type") or ""))
        cell.value = value
        if update.get("number_format"):
            cell.number_format = str(update.get("number_format"))
        applied.append({
            "sheet_name": sheet.title,
            "requested_cell": str(coordinate).upper(),
            "written_cell": cell.coordinate,
            "value": value,
        })
    return applied


def read_cell_number(sheet, coordinate, default=None):
    if not coordinate:
        return default
    cell = writable_cell_for_coordinate(sheet, coordinate)
    return number(cell.value, default)


def write_number(sheet, coordinate, value):
    cell = writable_cell_for_coordinate(sheet, coordinate)
    cell.value = value
    return cell.coordinate


def apply_row_calculations(wb, calculations, default_sheet_name):
    applied = []
    for item in calculations or []:
        if not isinstance(item, dict):
            continue
        sheet = sheet_for_update(wb, item.get("sheet_name") or default_sheet_name)
        base = read_cell_number(sheet, item.get("base_cell"))
        reading = read_cell_number(sheet, item.get("reading_cell"))
        multiplier = read_cell_number(sheet, item.get("multiplier_cell"), 1.0)
        unit_price = number(item.get("unit_price"))
        if base is None or reading is None or multiplier is None or unit_price is None:
            applied.append({"status": "skipped", "reason": "missing_number", "item": item})
            continue
        usage = round((reading - base) * multiplier, 2)
        amount = round(usage * unit_price, 2)
        usage_cell = write_number(sheet, item.get("usage_cell"), usage) if item.get("usage_cell") else ""
        price_cell = write_number(sheet, item.get("unit_price_cell"), unit_price) if item.get("unit_price_cell") else ""
        amount_cell = write_number(sheet, item.get("amount_cell"), amount) if item.get("amount_cell") else ""
        applied.append({
            "status": "applied",
            "sheet_name": sheet.title,
            "row_index": item.get("row_index"),
            "usage": usage,
            "amount": amount,
            "usage_cell": usage_cell,
            "unit_price_cell": price_cell,
            "amount_cell": amount_cell,
        })
    return applied


def cells_from_range(sheet, range_text):
    if not range_text:
        return []
    result = []
    for row in sheet[str(range_text)]:
        if isinstance(row, tuple):
            result.extend(row)
        else:
            result.append(row)
    return result


def apply_sum_updates(wb, updates, default_sheet_name):
    applied = []
    for update in updates or []:
        if not isinstance(update, dict) or not update.get("cell"):
            continue
        sheet = sheet_for_update(wb, update.get("sheet_name") or default_sheet_name)
        values = []
        for coordinate in update.get("cells") or []:
            value = read_cell_number(sheet, coordinate)
            if value is not None:
                values.append(value)
        for cell in cells_from_range(sheet, update.get("range")):
            value = number(cell.value)
            if value is not None:
                values.append(value)
        total = round(sum(values), 2)
        written_cell = write_number(sheet, update.get("cell"), total)
        applied.append({
            "sheet_name": sheet.title,
            "written_cell": written_cell,
            "value": total,
            "source_count": len(values),
        })
    return applied


def output_filename(args, input_meta):
    explicit_name = compact_text(args.get("output_filename"))
    if explicit_name:
        base = safe_filename_part(Path(explicit_name).stem)
        return (base or "property-utility-confirmation") + ".xlsx"
    source_name = safe_filename_part(Path(str(input_meta.get("filename") or "")).stem)
    payer = safe_filename_part(args.get("payer_name"))
    period = safe_filename_part(args.get("billing_period"))
    if "电费" in source_name:
        subject = "电费确认单"
    elif "水费" in source_name:
        subject = "水费确认单"
    else:
        subject = "水电确认单"
    parts = [part for part in [payer, subject, period] if part]
    if not parts:
        parts = [source_name or "property-utility-confirmation"]
    return "_".join(parts) + ".xlsx"


def set_label_value(sheet, labels, value):
    text_value = compact_text(value)
    if not text_value:
        return False
    normalized_labels = [normalize(label) for label in labels]
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            text = cell_text(cell)
            normalized = normalize(text)
            if not text or not any(label in normalized for label in normalized_labels):
                continue
            label = labels[0]
            if "：" in text:
                label = text.split("：", 1)[0]
            elif ":" in text:
                label = text.split(":", 1)[0]
            cell.value = f"{label}："
            writable_cell_to_right(sheet, row, col).value = text_value
            return True
    return False


def fill_total_row(sheet, cols, header_row, total_usage, total_amount):
    for row in range(header_row + 1, sheet.max_row + 1):
        row_text = "".join(cell_text(sheet.cell(row=row, column=col)) for col in range(1, min(sheet.max_column, 8) + 1))
        if "总计" not in row_text and "合计" not in row_text:
            continue
        if cols.get("usage"):
            sheet.cell(row=row, column=cols["usage"], value=total_usage)
        if cols.get("amount"):
            sheet.cell(row=row, column=cols["amount"], value=total_amount)
        return True
    return False


def fill_document_fields(sheet, args):
    payer = compact_text(args.get("payer_name"))
    period = compact_text(args.get("billing_period"))
    if payer or period:
        for row in range(1, min(sheet.max_row, 20) + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                text = cell_text(cell)
                normalized = normalize(text)
                if ("用电方名称" in text or "用水方名称" in text or "使用方名称" in text) and ("时间" in text or period):
                    prefix = "用电方名称" if "用电方名称" in text else ("用水方名称" if "用水方名称" in text else "使用方名称")
                    cell.value = f"{prefix}：{payer}    时间：{period}"
                    break
                if "用电方名称" in text or "用水方名称" in text or "使用方名称" in text:
                    prefix = "用电方名称" if "用电方名称" in text else ("用水方名称" if "用水方名称" in text else "使用方名称")
                    cell.value = f"{prefix}：{payer}"
                    break
                if "时间" in text and period and "用电方名称" not in text and "用水方名称" not in text:
                    cell.value = f"时间：{period}"
                    break
                if "用电方名称" in normalized or "用水方名称" in normalized or "使用方名称" in normalized:
                    cell.value = f"用电方名称：{payer}    时间：{period}"
                    break
    set_label_value(sheet, ["抄表"], args.get("meter_reader"))
    set_label_value(sheet, ["制表"], args.get("prepared_by"))


def find_header(sheet):
    aliases = {
        "item": ["单位名称", "客户名称", "公司名称", "租户名称", "用电位置", "用水位置", "项目", "类别"],
        "utility": ["费用类型", "水电类型", "类别"],
        "base": ["本月底数", "底数", "上月读数", "上期读数", "上月底数"],
        "reading": ["电表读数", "水表读数", "本期读数", "本月读数", "当前读数", "读数"],
        "multiplier": ["倍率", "倍数", "系数"],
        "usage": ["用电量", "用水量", "用量", "水电量"],
        "price": ["单价", "价格"],
        "amount": ["金额", "费用", "合计", "总计"],
    }
    best = None
    best_score = -1
    for row in range(1, min(sheet.max_row, 30) + 1):
        cols = {}
        for col in range(1, sheet.max_column + 1):
            text = normalize(cell_text(sheet.cell(row=row, column=col)))
            if not text:
                continue
            for key, names in aliases.items():
                if key in cols:
                    continue
                if any(normalize(alias) in text for alias in names):
                    cols[key] = col
        score = len(cols)
        if score > best_score:
            best = (row, cols)
            best_score = score
    if not best or "reading" not in best[1] or "base" not in best[1]:
        raise ValueError("template header not found: reading/base columns are required")
    row, cols = best
    next_col = sheet.max_column
    for key, label in [("usage", "用量"), ("price", "单价"), ("amount", "金额")]:
        if key not in cols:
            next_col += 1
            cols[key] = next_col
            sheet.cell(row=row, column=next_col, value=label)
    return row, cols


def number(value, default=None):
    if value is None or str(value).strip() == "":
        return default
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return default


def infer_utility(text):
    text = normalize(text)
    if "水" in text:
        return "water"
    if "电" in text or "kwh" in text:
        return "electricity"
    return ""


def first_present(record, keys):
    for key in keys:
        if key in record and record.get(key) not in (None, ""):
            return record.get(key)
    return None


def canonical_price_record(record):
    if not isinstance(record, dict):
        return None
    name = first_present(record, [
        "name",
        "item_name",
        "meter_name",
        "payer_name",
        "utility_name",
        "project",
        "名称",
        "项目",
        "表名",
        "收费对象",
        "用电方名称",
        "用水方名称",
    ])
    unit_price = first_present(record, [
        "unit_price",
        "price",
        "electricity_price",
        "water_price",
        "单价",
        "价格",
        "电价",
        "水价",
    ])
    if name is None or unit_price is None:
        return None
    unit_price = number(unit_price)
    if unit_price is None:
        return None
    utility_type = first_present(record, [
        "utility_type",
        "fee_type",
        "type",
        "category",
        "费用类型",
        "水电类型",
        "类型",
        "类别",
    ])
    utility_type = str(utility_type or "").strip().lower()
    if utility_type not in ("water", "electricity"):
        utility_type = infer_utility(utility_type or name)
    return {
        "name": str(name).strip(),
        "utility_type": utility_type,
        "unit_price": unit_price,
    }


def normalize_price_records(records):
    out = []
    if not isinstance(records, list):
        return out
    for record in records:
        item = canonical_price_record(record)
        if item:
            out.append(item)
    return out


def load_prices(args):
    with PRICE_FILE.open("r", encoding="utf-8") as f:
        prices = json.load(f)
    prices.extend(normalize_price_records(args.get("price_records")))
    prices.extend(normalize_price_records(args.get("unit_prices")))
    return prices


def match_price(item_name, utility_type, prices):
    norm_item = normalize(item_name)
    candidates = [p for p in prices if not utility_type or p.get("utility_type") == utility_type]
    best = None
    best_score = 0
    for price in candidates:
        norm_price = normalize(price.get("name"))
        score = 0
        if norm_item and norm_item in norm_price:
            score += 2
        if norm_price and norm_price in norm_item:
            score += 3
        if score > best_score:
            best = price
            best_score = score
    return best if best_score > 0 else None


def input_file_items(args):
    input_files = args.get("input_files") or {}
    confirmations = input_files.get("confirmations")
    if isinstance(confirmations, list) and confirmations:
        return confirmations
    confirmation = input_files.get("confirmation")
    if isinstance(confirmation, dict):
        return [confirmation]
    return []


def input_file_by_id(args):
    out = {}
    for item in input_file_items(args):
        if not isinstance(item, dict):
            continue
        file_id = compact_text(item.get("file_id"))
        if file_id:
            out[file_id] = item
    return out


def first_input_file(args):
    files = input_file_items(args)
    if not files:
        raise ValueError("input_files.confirmation or input_files.confirmations is required")
    first = files[0]
    if not isinstance(first, dict) or not first.get("path"):
        raise ValueError("input file path is required")
    return first


def inspect_input_files(args):
    files = input_file_items(args)
    if not files:
        raise ValueError("input_files.confirmation or input_files.confirmations is required")
    if len(files) == 1:
        wb = load_workbook(files[0]["path"])
        return {
            "success": True,
            "operation": "inspect",
            "workbook": inspect_workbook(wb, args),
        }
    workbooks = []
    for item in files:
        wb = load_workbook(item["path"])
        workbooks.append({
            "file_id": item.get("file_id"),
            "filename": item.get("filename"),
            "workbook": inspect_workbook(wb, args),
        })
    return {
        "success": True,
        "operation": "inspect",
        "workbook_count": len(workbooks),
        "workbooks": workbooks,
    }


def apply_edits_in_place(wb, args):
    sheet_name = args.get("sheet_name") or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet_name}")
    applied_cells = apply_cell_updates(wb, args.get("cell_updates"), sheet_name)
    applied_rows = apply_row_calculations(wb, args.get("row_calculations"), sheet_name)
    applied_sums = apply_sum_updates(wb, args.get("sum_updates"), sheet_name)
    if args.get("fill_basic_fields", True):
        fill_document_fields(wb[sheet_name], args)
    return {
        "success": True,
        "operation": "apply_edits",
        "sheet_name": sheet_name,
        "cell_update_count": len(applied_cells),
        "row_calculation_count": len([item for item in applied_rows if item.get("status") == "applied"]),
        "sum_update_count": len(applied_sums),
        "cell_updates": applied_cells,
        "row_calculations": applied_rows,
        "sum_updates": applied_sums,
    }


def save_workbook_artifact(wb, args, input_meta):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_name = output_filename(args, input_meta)
    output_path = OUTPUT_DIR / output_name
    wb.save(output_path)
    return output_name, output_path


def apply_edits_to_workbook(wb, args, input_meta):
    result = apply_edits_in_place(wb, args)
    output_name, output_path = save_workbook_artifact(wb, args, input_meta)
    result["filename"] = output_name
    result["artifacts"] = [{"path": str(output_path), "name": output_name}]
    return result


def batch_apply_edits(wb, args, input_meta):
    jobs = args.get("jobs")
    if not isinstance(jobs, list) or not jobs:
        raise ValueError("batch_apply_edits requires non-empty jobs")
    if len(jobs) > 10:
        raise ValueError("batch_apply_edits supports at most 10 jobs")

    results = []
    for index, job in enumerate(jobs, start=1):
        if not isinstance(job, dict):
            raise ValueError(f"batch job {index} must be an object")
        job_args = dict(args)
        job_args.update(job)
        job_args["input_files"] = args.get("input_files")
        job_args.pop("jobs", None)
        result = apply_edits_in_place(wb, job_args)
        result["job_index"] = index
        results.append(result)

    output_name, output_path = save_workbook_artifact(wb, args, input_meta)
    artifact = {"path": str(output_path), "name": output_name}
    return {
        "success": True,
        "operation": "batch_apply_edits",
        "job_count": len(results),
        "artifact_count": 1,
        "filename": output_name,
        "jobs": results,
        "artifacts": [artifact],
    }


def batch_files_apply_edits(args):
    file_jobs = args.get("file_jobs")
    if not isinstance(file_jobs, list) or not file_jobs:
        raise ValueError("batch_files_apply_edits requires non-empty file_jobs")
    if len(file_jobs) > 10:
        raise ValueError("batch_files_apply_edits supports at most 10 file_jobs")
    files_by_id = input_file_by_id(args)
    results = []
    artifacts = []
    for index, file_job in enumerate(file_jobs, start=1):
        if not isinstance(file_job, dict):
            raise ValueError(f"file job {index} must be an object")
        file_id = compact_text(file_job.get("confirmation_file_id"))
        input_meta = files_by_id.get(file_id)
        if not input_meta:
            raise ValueError(f"file job {index} confirmation_file_id not found in input_files.confirmations: {file_id}")
        wb = load_workbook(input_meta["path"])
        file_args = dict(args)
        file_args.update(file_job)
        file_args["input_files"] = args.get("input_files")
        file_args["jobs"] = file_job.get("jobs")
        file_args["operation"] = "batch_apply_edits"
        file_args.pop("file_jobs", None)
        result = batch_apply_edits(wb, file_args, input_meta)
        result["file_job_index"] = index
        result["confirmation_file_id"] = file_id
        results.append(result)
        artifacts.extend(result.get("artifacts") or [])
    return {
        "success": True,
        "operation": "batch_files_apply_edits",
        "file_job_count": len(results),
        "artifact_count": len(artifacts),
        "file_jobs": results,
        "artifacts": artifacts,
    }


def main():
    args = json.loads((sys.stdin.read() or "{}").lstrip("\ufeff"))
    operation = str(args.get("operation") or "").strip().lower()
    if operation in ("inspect", "preview"):
        print(json.dumps(inspect_input_files(args), ensure_ascii=False))
        return

    if operation in ("batch_files_apply_edits", "batch_files", "apply_files_batch"):
        print(json.dumps(batch_files_apply_edits(args), ensure_ascii=False))
        return

    input_meta = first_input_file(args)
    wb = load_workbook(input_meta["path"])
    sheet_name = args.get("sheet_name") or wb.sheetnames[0]

    if operation in ("batch_apply_edits", "batch", "apply_batch"):
        print(json.dumps(batch_apply_edits(wb, args, input_meta), ensure_ascii=False))
        return

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet_name}")

    if operation in ("apply_edits", "edit", "cell_updates") or args.get("cell_updates") or args.get("row_calculations") or args.get("sum_updates"):
        print(json.dumps(apply_edits_to_workbook(wb, args, input_meta), ensure_ascii=False))
        return

    sheet = wb[sheet_name]
    header_row, cols = find_header(sheet)
    fill_document_fields(sheet, args)
    prices = load_prices(args)

    matched = []
    unmatched = []
    total_usage = 0.0
    total_amount = 0.0
    for row in range(header_row + 1, sheet.max_row + 1):
        row_text = "".join(cell_text(sheet.cell(row=row, column=col)) for col in range(1, min(sheet.max_column, 8) + 1))
        if not row_text:
            continue
        if any(key in row_text for key in ["合计", "总计", "抄表", "制表"]):
            break
        item_name = cell_text(sheet.cell(row=row, column=cols.get("item", 1)))
        utility_type = infer_utility(cell_text(sheet.cell(row=row, column=cols.get("utility", cols.get("item", 1)))) or item_name)
        price = match_price(item_name, utility_type, prices)
        if not price:
            unmatched.append({"row_index": row, "item_name": item_name, "reason": "no price match"})
            continue
        base = number(sheet.cell(row=row, column=cols["base"]).value)
        reading = number(sheet.cell(row=row, column=cols["reading"]).value)
        multiplier = number(sheet.cell(row=row, column=cols.get("multiplier", 0)).value, 1.0) if cols.get("multiplier") else 1.0
        if base is None or reading is None:
            unmatched.append({"row_index": row, "item_name": item_name, "reason": "invalid reading/base"})
            continue
        usage = round((reading - base) * multiplier, 2)
        amount = round(usage * float(price["unit_price"]), 2)
        sheet.cell(row=row, column=cols["usage"], value=usage)
        sheet.cell(row=row, column=cols["price"], value=float(price["unit_price"]))
        sheet.cell(row=row, column=cols["amount"], value=amount)
        total_usage = round(total_usage + usage, 2)
        total_amount = round(total_amount + amount, 2)
        matched.append({"row_index": row, "item_name": item_name, "unit_price": price["unit_price"], "usage": usage, "amount": amount})
    total_filled = fill_total_row(sheet, cols, header_row, total_usage, total_amount)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_name = output_filename(args, input_meta)
    output_path = OUTPUT_DIR / output_name
    wb.save(output_path)
    print(json.dumps({
        "success": True,
        "filename": output_name,
        "matched_count": len(matched),
        "unmatched_count": len(unmatched),
        "total_usage": total_usage,
        "total_amount": total_amount,
        "total_filled": total_filled,
        "matched_items": matched,
        "unmatched_items": unmatched,
        "template_detection": {
            "sheet": sheet_name,
            "header_row": header_row,
            "columns": cols,
        },
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
